import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import json
import os
import requests
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- KONFIGURÁCIA ---
CONFIG_FILE = 'hospital_config.json'
HISTORY_FILE = 'room_history.json'
PRIVATE_CALENDAR_URL = "https://calendar.google.com/calendar/ical/fntnonk%40gmail.com/private-e8ce4e0639a626387fff827edd26b87f/basic.ics"

# ZMENA VERZIE NA v5 (Špeciálna logika pre Španika: Wolf + Malý disp.)
GIST_FILENAME_CONFIG = "hospital_config_v5.json"
GIST_FILENAME_HISTORY = "room_history_v5.json"

ROOMS_LIST = [
    (1, 3), (2, 3), (3, 3), (4, 3), (5, 3),
    (7, 1), (8, 3), (9, 3), (10, 1), (11, 1),
    (12, 2), (13, 2), (14, 2), (15, 2), (16, 2), (17, 2),
    (18, 3), (19, 3)
]

SENIOR_DOCTORS = ["Kurisova", "Vidulin", "Miklatkova"]

# --- GIST ULOŽISKO ---

def get_gist_id(filename):
    if "github" not in st.secrets:
        return None
    token = st.secrets["github"]["token"]
    headers = {"Authorization": f"token {token}"}
    try:
        resp = requests.get("https://api.github.com/gists", headers=headers)
        resp.raise_for_status()
        for gist in resp.json():
            if filename in gist['files']:
                return gist['id']
    except requests.exceptions.RequestException:
        pass
    return None

def load_data_from_gist(filename):
    if "github" not in st.secrets:
        return None
    gist_id = get_gist_id(filename)
    if not gist_id:
        return None
    try:
        token = st.secrets["github"]["token"]
        headers = {"Authorization": f"token {token}"}
        resp = requests.get(f"https://api.github.com/gists/{gist_id}", headers=headers)
        resp.raise_for_status()
        content = resp.json()['files'][filename]['content']
        return json.loads(content)
    except (requests.exceptions.RequestException, json.JSONDecodeError):
        return None

def save_data_to_gist(filename, data):
    if "github" not in st.secrets:
        return
    try:
        token = st.secrets["github"]["token"]
        headers = {"Authorization": f"token {token}"}
        gist_id = get_gist_id(filename)
        payload = {
            "description": f"Storage for {filename} (Streamlit App)",
            "public": False,
            "files": {
                filename: {"content": json.dumps(data, ensure_ascii=False, indent=2)}
            }
        }
        if gist_id:
            response = requests.patch(f"https://api.github.com/gists/{gist_id}", json=payload, headers=headers)
        else:
            response = requests.post("https://api.github.com/gists", json=payload, headers=headers)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        st.error(f"Chyba pri ukladaní do Gist ({filename}): {e}")

def _load_data(gist_filename, local_filename, default_factory):
    data = load_data_from_gist(gist_filename)
    if data is not None:
        return data
    if os.path.exists(local_filename):
        try:
            with open(local_filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (IOError, json.JSONDecodeError):
            pass
    return default_factory()

def load_config():
    config = _load_data(GIST_FILENAME_CONFIG, CONFIG_FILE, get_default_config)
    config, changed = migrate_homolova_to_vidulin(config)
    if 'closures' not in config:
        config['closures'] = {}
        changed = True
    if changed:
        save_config(config)
    return config

def load_history():
    return _load_data(GIST_FILENAME_HISTORY, HISTORY_FILE, lambda: {})

def save_config(config):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except IOError as e:
        st.error(f"Chyba pri lokálnom ukladaní configu: {e}")
    save_data_to_gist(GIST_FILENAME_CONFIG, config)

def save_history(history):
    try:
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except IOError as e:
        st.error(f"Chyba pri lokálnom ukladaní histórie: {e}")
    save_data_to_gist(GIST_FILENAME_HISTORY, history)

def get_default_config():
    return {
        "total_beds": 42,
        "closures": {},
        "ambulancie": {
            "Prijmova": {
                "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"],
                "priority": ["Kohutekova", "Kohutek", "Bystricky", "Zavrelova"]
            },
            "Velka dispenzarna": {
                "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"],
                "priority": ["Bocak", "Stratena", "Vidulin", "Kurisova",
                             "Blahova", "Hrabosova", "Miklatkova", "Martinka"]
            },
            "Mala dispenzarna": {
                "dni": ["Pondelok", "Piatok"],
                "priority": ["Spanik", "Stratena", "Vidulin", "Kurisova",
                             "Blahova", "Hrabosova", "Miklatkova"]
            },
            "Radio 2A": {
                "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"],
                "priority": ["Zavrelova", "Kohutek", "Kurisova",
                             "Miklatkova", "Bystricky"],
                "check_presence": ["Zavrelova", "Martinka"]
            },
            "Radio 2B": {
                "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"],
                "priority": ["Martinka"],
                "conditional_owner": "Martinka"
            },
            "Chemo 8A": {
                "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"],
                "priority": ["Hatalova", "Kohutek", "Stratena", "Bystricky"]
            },
            "Chemo 8B": {
                "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"],
                "priority": {
                    "0": ["Kohutek", "Stratena", "Bystricky", "Vidulin", "Blahova"],
                    "1": ["Kohutek", "Stratena", "Bystricky", "Vidulin", "Blahova"],
                    "default": ["Riedlova", "Kohutek", "Stratena",
                                "Bystricky", "Vidulin", "Blahova"]
                }
            },
            "Chemo 8C": {
                "dni": ["Utorok", "Streda", "Stvrtok"],
                "priority": ["Stratena", "Kohutek", "Bystricky", "Vidulin", "Blahova"]
            },
            "Wolf": {
                "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"],
                "priority": ["Spanik", "Miklatkova", "Kurisova", "Kohutek"]
            }
        },
        "lekari": {
            "Bystricky": {
                "moze": ["Prijmova", "Velka dispenzarna", "Mala dispenzarna",
                         "Radio 2A", "Chemo 8A", "Chemo 8B", "Chemo 8C", "Wolf"],
                "active": True
            },
            "Kohutek": {
                "moze": ["Oddelenie", "Prijmova", "Velka dispenzarna",
                         "Mala dispenzarna", "Radio 2A", "Chemo 8A",
                         "Chemo 8B", "Chemo 8C", "Wolf"],
                "pevne_dni": {"Pondelok": "Chemo 8B", "Utorok": "Chemo 8B"},
                "active": True
            },
            "Kohutekova": {
                "moze": ["Prijmova"],
                "pevne_dni": {
                    "Pondelok": "Prijmova", "Utorok": "Prijmova",
                    "Streda": "Prijmova", "Stvrtok": "Prijmova"
                },
                "nepracuje": ["Piatok"], "active": True
            },
            "Riedlova": {
                "moze": ["Chemo 8B"],
                "pevne_dni": {"Streda": "Chemo 8B", "Stvrtok": "Chemo 8B"},
                "nepracuje": ["Pondelok", "Utorok"], "active": True
            },
            "Zavrelova": {
                "moze": ["Radio 2A", "Prijmova"],
                "pevne_dni": {
                    "Pondelok": "Radio 2A", "Utorok": "Radio 2A", "Streda": "Radio 2A",
                    "Stvrtok": "Radio 2A", "Piatok": "Radio 2A"
                }, "active": True
            },
            "Martinka": {
                "moze": ["Radio 2B", "Oddelenie", "Velka dispenzarna"],
                "pevne_dni": {
                    "Pondelok": "Radio 2B", "Utorok": "Radio 2B", "Streda": "Radio 2B",
                    "Stvrtok": "Radio 2B", "Piatok": "Radio 2B"
                }, "active": True
            },
            "Hatalova": {
                "moze": ["Chemo 8A"],
                "pevne_dni": {
                    "Pondelok": "Chemo 8A", "Utorok": "Chemo 8A", "Streda": "Chemo 8A",
                    "Stvrtok": "Chemo 8A", "Piatok": "Chemo 8A"
                }, "active": True
            },
            "Stratena": {
                "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna",
                         "Chemo 8A", "Chemo 8B", "Chemo 8C"],
                "pevne_dni": {"Utorok": "Chemo 8C", "Streda": "Chemo 8C", "Stvrtok": "Chemo 8C"},
                "active": True
            },
            "Vidulin": {
                "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna",
                         "Chemo 8A", "Chemo 8B", "Chemo 8C"],
                "active": True
            },
            "Miklatkova": {"moze": ["Oddelenie", "Wolf"], "active": True},
            "Kurisova": {
                "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna", "Radio 2A", "Wolf"],
                "special": "veduca", "active": True
            },
            "Blahova": {
                "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna", "Chemo 8B", "Chemo 8C"],
                "active": False
            },
            "Hrabosova": {
                "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna"],
                "active": False
            },
            "Bocak": {
                "moze": ["Velka dispenzarna"],
                "pevne_dni": {
                    "Pondelok": "Velka dispenzarna", "Utorok": "Velka dispenzarna",
                    "Streda": "Velka dispenzarna", "Stvrtok": "Velka dispenzarna",
                    "Piatok": "Velka dispenzarna"
                }, "active": True
            },
            "Spanik": {
                "moze": ["Wolf", "Mala dispenzarna"],
                "pevne_dni": {
                    "Pondelok": "Mala dispenzarna", # V Po/Pi má primárne Malý disp., Wolf sa mu "prilepí" v kóde
                    "Utorok": "Wolf", 
                    "Streda": "Wolf", 
                    "Stvrtok": "Wolf",
                    "Piatok": "Mala dispenzarna"
                }, "active": True
            },
            "Kacurova": {"moze": ["Oddelenie"], "active": True},
            "Hunakova": {"moze": ["Oddelenie"], "active": True}
        }
    }

def migrate_homolova_to_vidulin(config):
    changed = False
    if "Homolova" in config["lekari"]:
        config["lekari"]["Vidulin"] = config["lekari"].pop("Homolova")
        changed = True
    for amb_name, amb_data in config["ambulancie"].items():
        if isinstance(amb_data["priority"], list):
            if "Homolova" in amb_data["priority"]:
                amb_data["priority"] = ["Vidulin" if x == "Homolova" else x for x in amb_data["priority"]]
                changed = True
        elif isinstance(amb_data["priority"], dict):
            for day_key, day_list in amb_data["priority"].items():
                if "Homolova" in day_list:
                    amb_data["priority"][day_key] = ["Vidulin" if x == "Homolova" else x for x in day_list]
                    changed = True
    return config, changed

# --- HIERARCHICKÉ ROZDEĽOVANIE IZIEB ---

def distribute_rooms(doctors_list, wolf_doc_name, previous_assignments=None, manual_core=None):
    if not doctors_list:
        return {}, {}
    if manual_core is None:
        manual_core = {}
    if previous_assignments is None:
        previous_assignments = {}

    head_doc = "Kurisova" if "Kurisova" in doctors_list else ("Miklatkova" if "Miklatkova" in doctors_list else None)
    deputy_doc = "Kohutek" if "Kohutek" in doctors_list else None
    rt_help_doc = "Miklatkova" if "Miklatkova" in doctors_list else None

    pure_workers = [d for d in doctors_list if d not in ["Kurisova", "Kohutek"]]
    if rt_help_doc and head_doc != rt_help_doc and rt_help_doc not in pure_workers:
        pure_workers.append(rt_help_doc)
    num_workers = len(pure_workers)

    caps = {d: 12 if d == wolf_doc_name else 15 for d in doctors_list}
    if deputy_doc: caps[deputy_doc] = 6
    if rt_help_doc:
        if num_workers >= 4: caps[rt_help_doc] = 9
        elif num_workers == 3: caps[rt_help_doc] = 12
        else: caps[rt_help_doc] = 15
    if num_workers < 2:
        for d in doctors_list: caps[d] = 15

    assignment = {d: [] for d in doctors_list}
    current_beds = {d: 0 for d in doctors_list}
    available_rooms = sorted(ROOMS_LIST, key=lambda x: x[0])

    for doc, nums in manual_core.items():
        if doc not in doctors_list: continue
        for num in nums:
            r_obj = next((r for r in available_rooms if r[0] == num), None)
            if not r_obj: continue
            assignment[doc].append(r_obj)
            current_beds[doc] += r_obj[1]
            available_rooms.remove(r_obj)

    active_assignees = [d for d in doctors_list]
    if num_workers >= 2 and head_doc in active_assignees and head_doc != rt_help_doc:
        active_assignees.remove(head_doc)

    if previous_assignments:
        divisors = len(active_assignees) if active_assignees else 1
        total_system_beds = sum(r[1] for r in ROOMS_LIST)
        ideal_load = total_system_beds / divisors
        threshold_base = ideal_load * 1.1

        for doc in active_assignees:
            if doc in previous_assignments and doc not in manual_core:
                my_hist_rooms = [r for r_num in previous_assignments[doc] if (r := next((room for room in ROOMS_LIST if room[0] == r_num), None)) and r in available_rooms]
                my_hist_rooms.sort(key=lambda x: x[0])
                is_senior = (doc in SENIOR_DOCTORS) or (doc == deputy_doc)
                
                temp_beds = sum(r[1] for r in my_hist_rooms)
                my_limit = caps.get(doc, 15)
                eff_limit = min(threshold_base, my_limit)

                while temp_beds > eff_limit and my_hist_rooms:
                    my_hist_rooms.pop() if is_senior else my_hist_rooms.pop(0)
                    temp_beds = sum(r[1] for r in my_hist_rooms)

                for r_obj in my_hist_rooms:
                    if current_beds[doc] + r_obj[1] <= caps.get(doc, 15):
                        assignment[doc].append(r_obj)
                        current_beds[doc] += r_obj[1]
                        available_rooms.remove(r_obj)

    while available_rooms:
        candidates = [d for d in active_assignees if current_beds[d] < caps.get(d, 15)]
        if not candidates:
            target_doc = head_doc if (head_doc and head_doc not in active_assignees) else (active_assignees[0] if active_assignees else None)
            if not target_doc: break
            candidates = [target_doc]

        candidates.sort(key=lambda w: current_beds[w])
        target_doc = candidates[0]
        
        doc_cap = caps.get(target_doc, 15)
        is_senior = (target_doc in SENIOR_DOCTORS)
        
        def room_score(r):
            deficit = doc_cap - current_beds[target_doc]
            fits = 1 if r[1] <= deficit else 0
            size_diff = abs(deficit - r[1])
            my_rooms = [x[0] for x in assignment[target_doc]]
            avg_pos = sum(my_rooms) / len(my_rooms) if my_rooms else 0
            dist = abs(r[0] - avg_pos) if my_rooms else (r[0] if is_senior else 20 - r[0])
            return (fits, -size_diff, -dist)
        
        available_rooms.sort(key=room_score, reverse=True)
        best_room = available_rooms.pop(0)

        assignment[target_doc].append(best_room)
        current_beds[target_doc] += best_room[1]

    result_text, result_raw = {}, {}
    for doc in doctors_list:
        rooms = sorted(assignment[doc], key=lambda x: x[0])
        result_raw[doc] = [r[0] for r in rooms]
        if not rooms:
            if doc == head_doc and num_workers >= 3:
                result_text[doc] = "RT oddelenie"
            else:
                result_text[doc] = "Wolf (0L)" if doc == wolf_doc_name else ""
        else:
            room_str = ", ".join([str(r[0]) for r in rooms])
            suffix = " + Wolf" if doc == wolf_doc_name else (" + RT oddelenie" if doc == head_doc else "")
            result_text[doc] = f"{room_str}{suffix}"
    return result_text, result_raw

# --- DATA PROCESSING ---

def _unfold_ical_lines(ical_text):
    unfolded = []
    for line in ical_text.splitlines():
        if line.startswith((" ", "\t")) and unfolded:
            unfolded[-1] += line[1:]
        else:
            unfolded.append(line)
    return unfolded

def _parse_ical_datetime(raw_value):
    value = raw_value.strip()
    formats = (
        "%Y%m%d",
        "%Y%m%dT%H%M%S",
        "%Y%m%dT%H%M",
        "%Y%m%dT%H%M%SZ",
        "%Y%m%dT%H%MZ",
    )
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None

def _iter_ical_events(ical_text):
    event = None
    for line in _unfold_ical_lines(ical_text):
        if line == "BEGIN:VEVENT":
            event = {}
            continue
        if line == "END:VEVENT":
            if event:
                yield event
            event = None
            continue
        if event is None or ":" not in line:
            continue

        key, value = line.split(":", 1)
        prop_name = key.split(";", 1)[0].upper()
        if prop_name in {"SUMMARY", "DTSTART", "DTEND"}:
            event[prop_name] = value.strip()

def get_ical_events(start_date, end_date):
    try:
        response = requests.get(PRIVATE_CALENDAR_URL)
        response.raise_for_status()
        absences = {}
        for event in _iter_ical_events(response.text):
            dt_start = _parse_ical_datetime(event.get("DTSTART", ""))
            dt_end = _parse_ical_datetime(event.get("DTEND", ""))
            raw = event.get("SUMMARY", "").strip()
            if not dt_start or not dt_end or not raw:
                continue

            ev_start, ev_end = dt_start.date(), dt_end.date()
            if ev_end < start_date.date() or ev_start > end_date.date():
                continue
            name, typ = raw, "Dovolenka"

            if raw.upper().endswith('PN'):
                typ, name = "PN", raw[:-2].rstrip(' -')
            elif raw.upper().endswith('VZ'):
                typ, name = "Vzdelávanie", raw[:-2].rstrip(' -')
            elif raw.upper().endswith('S') and not raw.upper().endswith('OS'):
                typ, name = "Stáž", raw[:-1].rstrip(' -')
            elif '-' in raw and typ == "Dovolenka":
                parts = raw.split('-')
                name = parts[0].strip()
                suffix = parts[1].strip().upper() if len(parts) > 1 else ""
                if suffix == 'S': typ = "Stáž"
                elif suffix == 'PN': typ = "PN"
                elif suffix == 'VZ': typ = "Vzdelávanie"
            
            curr, limit = max(ev_start, start_date.date()), min(ev_end, end_date.date())
            while curr < limit:
                absences.setdefault(curr.strftime('%Y-%m-%d'), {})[name] = typ
                curr += timedelta(days=1)
        return absences
    except requests.RequestException as e:
        st.error(f"Chyba pri sťahovaní kalendára: {e}")
        return {}

def generate_data_structure(config, absences, start_date):
    days_map = {0: "Pondelok", 1: "Utorok", 2: "Streda", 3: "Stvrtok", 4: "Piatok"}
    weekday = start_date.weekday()
    thursday = start_date + timedelta(days=(3 - weekday) % 7)
    
    dates, data_grid = [], {}
    all_doctors = sorted([d for d, props in config['lekari'].items() if props.get('active')])

    history = load_history()
    day_before_str = (thursday - timedelta(days=1)).strftime('%Y-%m-%d')
    last_day_assignments = history.get(day_before_str, {})
    manual_all = st.session_state.get("manual_core", {})
    
    closures = config.get('closures', {})
    
    for i in range(7):
        curr_date = thursday + timedelta(days=i)
        day_name = days_map.get(curr_date.weekday())
        if not day_name: continue

        date_str = curr_date.strftime('%d.%m.%Y')
        date_key = curr_date.strftime('%Y-%m-%d')
        dates.append(date_str)
        
        day_absences = absences.get(date_key, {})
        closed_today = closures.get(date_key, [])
        
        data_grid[date_str] = {}

        available = [d for d in all_doctors if d not in day_absences and day_name not in config['lekari'][d].get('nepracuje', [])]
        assigned_amb = {}

        # 1. Pevné dni (najvyššia priorita)
        for doc in list(available):
            if fixed := config['lekari'][doc].get('pevne_dni', {}).get(day_name):
                for t in [t.strip() for t in fixed.split(',')]:
                    if t in closed_today:
                        assigned_amb[t] = "ZATVORENÉ"
                    else:
                        assigned_amb[t] = doc
                available.remove(doc)

        # 2. Poradie obsadzovania ambulancií
        processing_order = ["Radio 2A", "Radio 2B", "Chemo 8B", "Chemo 8A", "Chemo 8C", "Wolf", "Prijmova", "Velka dispenzarna", "Mala dispenzarna"]

        for amb_name in processing_order:
            # ŠPECIÁLNA LOGIKA PRE WOLF A ŠPÁNIKA:
            # Ak je to Wolf a Španik je už priradený na Malú dispenzárnu (z pevných dní),
            # tak mu priradíme AJ Wolf.
            if amb_name == "Wolf":
                if "Spanik" in all_doctors and "Spanik" not in day_absences:
                    # Ak má Španik už Malú dispenzárnu (Po/Pi), priradíme mu aj Wolf
                    if assigned_amb.get("Mala dispenzarna") == "Spanik":
                        assigned_amb["Wolf"] = "Spanik"
                        continue # Wolf vybavený, ideme ďalej
                    
                    # Ak je voľný (Ut-Štv), vezme ho cez pevné dni (už spracované vyššie), 
                    # alebo cez prioritu nižšie.

            if amb_name in assigned_amb: 
                continue
            
            if amb_name in closed_today:
                assigned_amb[amb_name] = "ZATVORENÉ"
                continue

            amb_info = config['ambulancie'][amb_name]
            if day_name not in amb_info['dni']:
                assigned_amb[amb_name] = "---"
                continue
            
            if amb_name == "Radio 2B" and "Martinka" not in available:
                assigned_amb[amb_name] = "ZATVORENÉ"
                continue

            prio_list = amb_info['priority']
            if isinstance(prio_list, dict):
                prio_list = prio_list.get(str(curr_date.weekday()), prio_list.get('default', []))

            for doc in prio_list:
                if doc in available and amb_name in config['lekari'][doc].get('moze', []):
                    assigned_amb[amb_name] = doc
                    available.remove(doc)
                    break
            if amb_name not in assigned_amb:
                assigned_amb[amb_name] = "NEOBSADENÉ"
        
        for amb, val in assigned_amb.items(): 
            data_grid[date_str][amb] = val

        # 3. Izby a Wolf
        wolf_doc = assigned_amb.get("Wolf")
        
        if "ODDELENIE (Celé)" in closed_today:
            room_text_map, room_raw_map = {}, {}
            for doc in all_doctors:
                if doc in day_absences:
                    continue
                # Ak už má ambulanciu (vrátane Wolf/Mala disp), preskočíme
                if doc in assigned_amb.values():
                    continue
                if "Oddelenie" in config['lekari'][doc].get('moze', []):
                    room_text_map[doc] = "ZATVORENÉ"
        else:
            ward_candidates = [d for d in available if "Oddelenie" in config['lekari'][d].get('moze', [])]
            if wolf_doc and wolf_doc not in ward_candidates and "Oddelenie" in config['lekari'].get(wolf_doc, {}).get('moze', []):
                ward_candidates.append(wolf_doc)
            
            manual_for_day = manual_all.get(start_date.strftime('%Y-%m-%d'), {})
            room_text_map, room_raw_map = distribute_rooms(ward_candidates, wolf_doc, last_day_assignments, manual_for_day)
            
            last_day_assignments = room_raw_map
            history[date_key] = room_raw_map

        # 4. Finálne pridelenie
        for doc in all_doctors:
            if doc in day_absences: 
                data_grid[date_str][doc] = day_absences[doc]
            elif doc in room_text_map: 
                data_grid[date_str][doc] = room_text_map[doc]
            else: 
                # Zobrazíme všetky ambulancie, ktoré má lekár v tento deň
                my_ambs = [a for a, d in assigned_amb.items() if d == doc]
                if my_ambs:
                    data_grid[date_str][doc] = " + ".join(my_ambs)
                else:
                    data_grid[date_str][doc] = ""

    save_history(history)
    return dates, data_grid, all_doctors

# --- VIZUALIZÁCIA ---

def create_display_df(dates, data_grid, all_doctors, motto, config):
    rows = []
    ward_doctors = [d for d in all_doctors if "Oddelenie" in config['lekari'][d].get('moze', [])]
    
    display_map = {
        "Radio 2A": "RT ambulancia",
        "Velka dispenzarna": "veľký dispenzár",
        "Mala dispenzarna": "malý dispenzár"
    }

    rows.append(["Oddelenie"] + dates)
    for doc in ward_doctors:
        vals = []
        for date in dates:
            val = data_grid[date].get(doc, "")
            for old, new in display_map.items():
                val = val.replace(old, new)
            vals.append(val)
        rows.append([f"Dr {doc}"] + vals)
    rows.append([motto or "Motto"] + [""] * len(dates))

    sections = [
        ("Konziliárna amb", ["Prijmova"]),
        ("RT ambulancie", ["Radio 2A", "Radio 2B"]),
        ("Chemo amb", ["Chemo 8A", "Chemo 8B", "Chemo 8C"]),
        ("Disp. Ambulancia", ["Velka dispenzarna", "Mala dispenzarna"]),
        ("RTG Terapia", ["Wolf"])
    ]
    
    for title, amb_list in sections:
        rows.append([title] + dates)
        for amb in amb_list:
            display_name = display_map.get(amb, amb)
            vals = []
            for date in dates:
                val = data_grid[date].get(amb, "")
                val = val.replace("---", "").replace("NEOBSADENÉ", "???")
                vals.append(val)
            rows.append([display_name] + vals)
        rows.append([""] * (len(dates) + 1))
        
    return pd.DataFrame(rows)

def create_excel_report(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Rozpis")
        ws = writer.sheets['Rozpis']
        
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.cell(row=1, column=1, value=f"Rozpis prác Onkologická klinika {df.columns[1]} - {df.columns[-1]}").font = bold_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws['A1'].alignment = center_align

        for r_idx, row in enumerate(df.iterrows(), 2):
            is_header = row[1][0] in ["Oddelenie", "Konziliárna amb", "RT ambulancie", "Chemo amb", "Disp. Ambulancia", "RTG Terapia"]
            is_motto = (row[1][0] == st.session_state.get('motto', 'Motto'))
            
            for c_idx, value in enumerate(row[1], 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_align
                if is_header or (c_idx==1 and not is_motto): 
                    cell.font = bold_font
                if is_motto:
                    ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=len(df.columns))
                    cell.font = Font(bold=True, italic=True)
                    cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                    break
        
        ws.column_dimensions['A'].width = 25
        for i in range(2, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
            
    return output.getvalue()

# --- UI ---

st.set_page_config(page_title="Rozpis FN Trenčín", layout="wide")
st.title("🏥 Rozpis prác - Onkologická klinika FN Trenčín")

if 'config' not in st.session_state:
    st.session_state.config = load_config()
if 'manual_core' not in st.session_state:
    st.session_state.manual_core = {}

mode = st.sidebar.radio("Navigácia", ["🚀 Generovať rozpis", "⚙️ Nastavenia lekárov", "🏥 Nastavenia ambulancií"])

if mode == "🚀 Generovať rozpis":
    c1, c2 = st.columns([2, 2])
    st.session_state.motto = c1.text_input("📢 Motto týždňa (nepovinné):", placeholder="Sem napíšte motto...")
    start_d = c2.date_input("Začiatok rozpisu (vypočíta najbližší štvrtok):", datetime.now())

    with st.expander("📅 Výnimky a zatváranie ambulancií (Manuálne)"):
        st.info("Vyberte rozsah dátumov (napr. Vianoce) a čo má byť zatvorené. Ak necháte výber prázdny a uložíte, výnimky sa pre dané dni zrušia.")
        c_ex1, c_ex2, c_ex3 = st.columns([1, 2, 1])
        
        d_range = c_ex1.date_input(
            "Rozsah dátumov (Od - Do):",
            value=[],
            help="Kliknite najprv na dátum začiatku, potom na dátum konca."
        )
        amb_options = ["ODDELENIE (Celé)"] + list(st.session_state.config['ambulancie'].keys())
        selected_closures = c_ex2.multiselect(
            "Čo má byť v tomto období ZATVORENÉ?",
            options=amb_options
        )
        
        if c_ex3.button("💾 Uložiť pre celé obdobie"):
            if 'closures' not in st.session_state.config:
                st.session_state.config['closures'] = {}
            
            if len(d_range) == 2:
                start_r, end_r = d_range
                curr = start_r
                cnt = 0
                while curr <= end_r:
                    d_key = curr.strftime('%Y-%m-%d')
                    if selected_closures:
                        st.session_state.config['closures'][d_key] = selected_closures
                    else:
                        if d_key in st.session_state.config['closures']:
                            del st.session_state.config['closures'][d_key]
                    curr += timedelta(days=1)
                    cnt += 1
                save_config(st.session_state.config)
                st.success(f"Nastavenia aplikované na {cnt} dní.")
            elif len(d_range) == 1:
                d_key = d_range[0].strftime('%Y-%m-%d')
                if selected_closures:
                    st.session_state.config['closures'][d_key] = selected_closures
                else:
                    if d_key in st.session_state.config['closures']:
                        del st.session_state.config['closures'][d_key]
                save_config(st.session_state.config)
                st.success("Nastavené pre 1 deň.")
            else:
                st.warning("Prosím, vyberte dátum.")

    st.markdown("### Manuálne pridelenie izieb")
    manual_core_input = {}
    ward_docs = [d for d, p in st.session_state.config["lekari"].items() if "Oddelenie" in p.get("moze", []) and p.get("active")]
    cols = st.columns(2)
    for i, doc in enumerate(ward_docs):
        txt = cols[i % 2].text_input(f"Dr {doc} – izby (oddelené čiarkou)", key=f"core_{doc}")
        if txt.strip(): 
            manual_core_input[doc] = [int(p.strip()) for p in txt.split(',') if p.strip().isdigit()]
    
    if manual_core_input: 
        st.session_state.manual_core[start_d.strftime('%Y-%m-%d')] = manual_core_input

    c3, c4 = st.columns([1, 1])
    if c3.button("🚀 Generovať nový rozpis", type="primary"):
        with st.spinner("Sťahujem kalendár a počítam..."):
            end_d = start_d + timedelta(days=14)
            absences = get_ical_events(datetime.combine(start_d, datetime.min.time()), datetime.combine(end_d, datetime.min.time()))
            dates, grid, docs = generate_data_structure(st.session_state.config, absences, start_d)
            
            df_display = create_display_df(dates, grid, docs, st.session_state.motto, st.session_state.config)
            df_display.columns = ["Sekcia / Dátum"] + dates
            st.session_state.df_display = df_display

            st.success("✅ Hotovo! Izby sú synchronizované s históriou.")
    
    if c4.button("🗑️ Vymazať históriu izieb"):
        save_history({})
        if os.path.exists(HISTORY_FILE): 
            os.remove(HISTORY_FILE)
        st.success("História vymazaná a odoslaná na Gist.")

    if 'df_display' in st.session_state:
        df_for_excel = st.session_state.df_display.copy()
        df_for_excel.iloc[0, 1:] = df_for_excel.columns[1:] 
        xlsx_data = create_excel_report(df_for_excel)
        st.download_button(label="⬇️ Stiahnuť EXCEL Rozpis (.xlsx)", data=xlsx_data, file_name=f"Rozpis.xlsx")
        st.markdown("---")
        st.subheader("📄 Náhľad")
        st.dataframe(st.session_state.df_display, use_container_width=True, hide_index=True)

elif mode == "⚙️ Nastavenia lekárov":
    st.header("Správa lekárov")
    col_new, col_btn = st.columns([3, 1])
    new_doc = col_new.text_input("Pridať nového lekára (meno)")
    if col_btn.button("➕ Pridať") and new_doc:
        if new_doc not in st.session_state.config['lekari']:
            st.session_state.config['lekari'][new_doc] = {"moze": ["Oddelenie"], "active": True}
            save_config(st.session_state.config)
            st.success(f"Lekár {new_doc} pridaný")
            st.rerun()

    for doc, data in st.session_state.config['lekari'].items():
        with st.expander(f"{doc} {'(Neaktívny)' if not data.get('active', True) else ''}"):
            c1, c2 = st.columns(2)
            act = c1.checkbox("Aktívny", value=data.get('active', True), key=f"act_{doc}")
            all_places = list(st.session_state.config['ambulancie'].keys()) + ["Oddelenie"]
            can_do = st.multiselect("Môže pracovať na:", all_places, default=[p for p in data.get('moze', []) if p in all_places], key=f"can_{doc}")
            
            if act != data.get('active', True) or can_do != data.get('moze', []):
                data['active'] = act
                data['moze'] = can_do
                save_config(st.session_state.config)
                st.rerun()
            
            if st.button(f"🗑️ Odstrániť {doc}", key=f"del_{doc}"):
                del st.session_state.config['lekari'][doc]
                save_config(st.session_state.config)
                st.rerun()
    
elif mode == "🏥 Nastavenia ambulancií":
    st.header("Priority ambulancií")
    ambs = st.session_state.config['ambulancie']
    sel_amb = st.selectbox("Vyberte ambulanciu na úpravu", list(ambs.keys()))
    curr_amb = ambs[sel_amb]
    st.info(f"Dni prevádzky: {', '.join(curr_amb['dni'])}")
    prio = curr_amb['priority']
    
    if isinstance(prio, list):
        new_prio_text = st.text_area(f"Zoznam priorít pre {sel_amb} (oddelené čiarkou)", ", ".join(prio))
        if st.button("💾 Uložiť priority"):
            ambs[sel_amb]['priority'] = [x.strip() for x in new_prio_text.split(',')]
            save_config(st.session_state.config)
            st.success("Priority aktualizované")
    else:
        st.warning("Táto ambulancia má komplexné priority (podľa dní). Upravte ich priamo v JSON/Gist.")
